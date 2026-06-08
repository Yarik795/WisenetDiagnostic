#!/usr/bin/env python3
"""Экспорт списка регистраторов в Excel: модель, IP, объект, S/N, дата производства.

Данные:
  - config.json — объект, IP, имя регистратора;
  - monitoring.db (recorder_metrics) — модель, серийный номер, дата производства.

Запуск (из корня репозитория):
  python scripts/export_recorders_serial.py
  python scripts/export_recorders_serial.py -o report.xlsx
  python scripts/export_recorders_serial.py --config config.json --db data/monitoring.db

Переменные окружения:
  CONFIG_PATH   — путь к config.json
  STATE_DB_PATH — путь к monitoring.db
"""

from __future__ import annotations

import argparse
import os
import sys
from datetime import datetime
from pathlib import Path
from typing import Any, Optional

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
BACKEND = ROOT / "backend"
if str(BACKEND) not in sys.path:
    sys.path.insert(0, str(BACKEND))

from app.config_store import ConfigStore  # noqa: E402
from app.models import Recorder  # noqa: E402
from app.state_store import StateStore  # noqa: E402
from app.ui.metrics_helpers import format_manufacture_date  # noqa: E402

DEFAULT_CONFIG_PATH = ROOT / "config.json"
DEFAULT_DB_PATH = ROOT / "data" / "monitoring.db"
DEFAULT_OUTPUT_DIR = ROOT / "data"

HEADERS = (
    "Наименование объекта",
    "Модель",
    "IP-адрес",
    "Серийный номер",
    "Дата производства",
)

NO_SERIAL_TEXT = "нет серийного номера"
NO_DATE_TEXT = "—"


def _default_output_path() -> Path:
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return DEFAULT_OUTPUT_DIR / f"recorders_serial_{stamp}.xlsx"


def _recorder_ip(recorder: Recorder) -> str:
    if recorder.port in (80, 443):
        return recorder.host
    return f"{recorder.host}:{recorder.port}"


def _model_value(recorder: Recorder, metrics_model: Optional[str]) -> str:
    if metrics_model and metrics_model.strip():
        return metrics_model.strip()
    if recorder.name and recorder.name.strip():
        return recorder.name.strip()
    return "—"


def _serial_value(raw: Optional[str]) -> str:
    if raw and raw.strip():
        return raw.strip()
    return NO_SERIAL_TEXT


def _manufacture_date_value(raw: Optional[str], *, has_serial: bool) -> str:
    if raw and raw.strip():
        return format_manufacture_date(raw)
    if has_serial:
        return "не определена"
    return NO_DATE_TEXT


def build_rows(
    recorders: list[Recorder],
    metrics_by_id: dict[str, Any],
) -> list[tuple[str, str, str, str, str]]:
    rows: list[tuple[str, str, str, str, str]] = []
    for recorder in sorted(recorders, key=lambda r: (r.object_name.lower(), r.host)):
        metrics = metrics_by_id.get(recorder.id)
        model = _model_value(recorder, metrics.model if metrics else None)
        serial_raw = metrics.serial_number if metrics else None
        serial = _serial_value(serial_raw)
        has_serial = serial != NO_SERIAL_TEXT
        mfg_raw = metrics.manufacture_date if metrics else None
        mfg = _manufacture_date_value(mfg_raw, has_serial=has_serial)
        rows.append(
            (
                recorder.object_name,
                model,
                _recorder_ip(recorder),
                serial,
                mfg,
            )
        )
    return rows


def write_xlsx(path: Path, rows: list[tuple[str, str, str, str, str]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Регистраторы"

    header_font = Font(bold=True)
    for col, title in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.font = header_font

    for row_idx, row in enumerate(rows, start=2):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    for col in range(1, len(HEADERS) + 1):
        letter = get_column_letter(col)
        max_len = max(
            len(str(ws.cell(row=r, column=col).value or "")) for r in range(1, ws.max_row + 1)
        )
        ws.column_dimensions[letter].width = min(max(max_len + 2, 12), 60)

    ws.freeze_panes = "A2"
    wb.save(path)


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Экспорт регистраторов (модель, IP, объект, S/N, дата производства) в Excel"
    )
    parser.add_argument(
        "--config",
        type=Path,
        default=Path(os.environ.get("CONFIG_PATH", DEFAULT_CONFIG_PATH)),
        help="Путь к config.json",
    )
    parser.add_argument(
        "--db",
        type=Path,
        default=Path(os.environ.get("STATE_DB_PATH", DEFAULT_DB_PATH)),
        help="Путь к monitoring.db",
    )
    parser.add_argument(
        "-o",
        "--output",
        type=Path,
        default=None,
        help="Файл .xlsx (по умолчанию: data/recorders_serial_YYYYMMDD_HHMMSS.xlsx)",
    )
    args = parser.parse_args()

    config_path = args.config.resolve()
    db_path = args.db.resolve()
    output_path = (args.output or _default_output_path()).resolve()

    if not config_path.is_file():
        print(f"Ошибка: config не найден: {config_path}", file=sys.stderr)
        return 1
    if not db_path.is_file():
        print(f"Ошибка: база не найдена: {db_path}", file=sys.stderr)
        return 1

    store = ConfigStore(config_path)
    state = StateStore(db_path)
    config = store.load()
    metrics_by_id = {m.recorder_id: m for m in state.list_recorder_metrics()}

    rows = build_rows(config.recorders, metrics_by_id)
    write_xlsx(output_path, rows)

    with_serial = sum(1 for r in rows if r[3] != NO_SERIAL_TEXT)
    print(f"Записано регистраторов: {len(rows)}")
    print(f"С серийным номером: {with_serial}")
    print(f"Без серийного номера: {len(rows) - with_serial}")
    print(f"Файл: {output_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
