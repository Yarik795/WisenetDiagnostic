"""Чтение устройств из CMDB (Excel) для синхронизации config.json."""

from __future__ import annotations

import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Literal, Optional

FUNCTIONAL_TYPE_VIDEO = "Видеорегистраторы"
COL_IP = "IP"
COL_ADDRESS = "Адрес"
COL_MODEL = "Модель устройства"
COL_FUNCTIONAL_TYPE = "Функциональный тип"
COL_MANUFACTURER = "Производитель устройства"
COL_MAC = "MAC"

MANUFACTURER_SKUD = "Бастион"
MANUFACTURER_BIO = "PocketKey"

CmdbDeviceKind = Literal["tsv", "skud", "bio"]

_REQUIRED_HEADERS = {COL_IP, COL_ADDRESS, COL_MODEL, COL_MANUFACTURER, COL_MAC}

# Latin letters that look like Cyrillic (common CMDB export corruption).
_HOMOGLYPH_MAP = str.maketrans(
    {
        "A": "А",
        "a": "а",
        "B": "В",
        "C": "С",
        "c": "с",
        "E": "Е",
        "e": "е",
        "H": "Н",
        "K": "К",
        "M": "М",
        "O": "О",
        "o": "о",
        "P": "Р",
        "p": "р",
        "T": "Т",
        "X": "Х",
        "x": "х",
        "Y": "У",
        "y": "у",
    }
)


def normalize_homoglyphs(text: str) -> str:
    if not any("\u0400" <= ch <= "\u04FF" for ch in text):
        return text
    return text.translate(_HOMOGLYPH_MAP)


def normalize_cmdb_label(value: Any) -> str:
    """Normalize CMDB headers and categorical fields (not IP/MAC/host)."""
    if value is None:
        return ""
    text = re.sub(r"\s+", " ", str(value).strip())
    return normalize_homoglyphs(text)


def normalize_header(value: Any) -> str:
    return normalize_cmdb_label(value)


def _cell_str(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def find_header_row(rows: list[list[Any]]) -> int:
    for idx, row in enumerate(rows):
        headers = {normalize_header(c) for c in row if c is not None and str(c).strip()}
        has_type_or_vendor = COL_FUNCTIONAL_TYPE in headers or COL_MANUFACTURER in headers
        if COL_IP in headers and has_type_or_vendor:
            return idx
    raise ValueError(
        f"Строка заголовков не найдена (нужны колонки «{COL_IP}» и "
        f"«{COL_FUNCTIONAL_TYPE}» или «{COL_MANUFACTURER}»)"
    )


def build_col_index(header_row: list[Any]) -> dict[str, int]:
    index: dict[str, int] = {}
    for col, cell in enumerate(header_row):
        name = normalize_header(cell)
        if name and name not in index:
            index[name] = col
    missing = _REQUIRED_HEADERS - set(index)
    if missing:
        raise ValueError(f"В заголовке отсутствуют колонки: {', '.join(sorted(missing))}")
    return index


def classify_cmdb_row(func_type: str, manufacturer: str) -> Optional[CmdbDeviceKind]:
    if func_type == FUNCTIONAL_TYPE_VIDEO:
        return "tsv"
    if manufacturer == MANUFACTURER_SKUD:
        return "skud"
    if manufacturer == MANUFACTURER_BIO:
        return "bio"
    return None


@dataclass(frozen=True)
class CmdbDeviceRow:
    host: str
    object_name: str
    name: Optional[str]
    mac: Optional[str]
    device_kind: CmdbDeviceKind
    source_row: int


# Backward-compatible alias
CmdbRecorderRow = CmdbDeviceRow


@dataclass
class CmdbParseResult:
    rows: list[CmdbDeviceRow]
    total_data_rows: int
    skipped_empty_ip: int
    skipped_unclassified: int
    counts_by_kind: dict[str, int] = field(default_factory=dict)

    @property
    def skipped_wrong_type(self) -> int:
        return self.skipped_unclassified


def parse_cmdb_grid(rows: list[list[Any]]) -> CmdbParseResult:
    header_idx = find_header_row(rows)
    col_index = build_col_index(rows[header_idx])
    ip_col = col_index[COL_IP]
    addr_col = col_index[COL_ADDRESS]
    model_col = col_index[COL_MODEL]
    mac_col = col_index[COL_MAC]
    vendor_col = col_index[COL_MANUFACTURER]
    type_col = col_index.get(COL_FUNCTIONAL_TYPE)

    result_rows: list[CmdbDeviceRow] = []
    skipped_empty_ip = 0
    skipped_unclassified = 0
    counts_by_kind: dict[str, int] = {"tsv": 0, "skud": 0, "bio": 0}
    total_data_rows = 0

    for row_idx, row in enumerate(rows[header_idx + 1 :], start=header_idx + 2):
        if not row or all(c is None or str(c).strip() == "" for c in row):
            continue
        total_data_rows += 1

        host = _cell_str(row[ip_col] if ip_col < len(row) else None)
        if not host:
            skipped_empty_ip += 1
            continue

        func_type = ""
        if type_col is not None:
            func_type = normalize_cmdb_label(
                row[type_col] if type_col < len(row) else None
            )
        manufacturer = normalize_cmdb_label(
            row[vendor_col] if vendor_col < len(row) else None
        )
        device_kind = classify_cmdb_row(func_type, manufacturer)
        if device_kind is None:
            skipped_unclassified += 1
            continue

        object_name = _cell_str(row[addr_col] if addr_col < len(row) else None)
        model = _cell_str(row[model_col] if model_col < len(row) else None)
        mac_raw = _cell_str(row[mac_col] if mac_col < len(row) else None)
        result_rows.append(
            CmdbDeviceRow(
                host=host,
                object_name=object_name,
                name=model or None,
                mac=mac_raw or None,
                device_kind=device_kind,
                source_row=row_idx,
            )
        )
        counts_by_kind[device_kind] = counts_by_kind.get(device_kind, 0) + 1

    return CmdbParseResult(
        rows=result_rows,
        total_data_rows=total_data_rows,
        skipped_empty_ip=skipped_empty_ip,
        skipped_unclassified=skipped_unclassified,
        counts_by_kind=counts_by_kind,
    )


def read_cmdb_xlsx(path: Path) -> CmdbParseResult:
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb.active
        grid: list[list[Any]] = []
        for row in ws.iter_rows(values_only=True):
            grid.append(list(row))
    finally:
        wb.close()
    return parse_cmdb_grid(grid)


# --- merge with existing config ---


def new_device_id(device_kind: CmdbDeviceKind) -> str:
    import uuid

    prefix = {"tsv": "nvr", "skud": "skud", "bio": "bio"}[device_kind]
    return f"{prefix}-{uuid.uuid4().hex[:8]}"


def new_recorder_id() -> str:
    return new_device_id("tsv")


@dataclass
class MergeStats:
    preserved: int
    added: int
    removed: int


@dataclass
class MergeError:
    source_row: int
    message: str


def _recorder_kind(recorder: Any) -> str:
    return getattr(recorder, "device_kind", None) or "tsv"


def _merge_key(host: str, device_kind: str) -> tuple[str, str]:
    return (host, device_kind)


def merge_devices_from_cmdb(
    cmdb_rows: list[CmdbDeviceRow],
    existing: list[Any],
) -> tuple[list[Any], MergeStats, list[MergeError]]:
    from pydantic import ValidationError

    from app.device_kinds import CMDB_MANAGED_KINDS
    from app.models import Recorder

    preserved_non_cmdb = [
        r for r in existing if _recorder_kind(r) not in CMDB_MANAGED_KINDS
    ]
    remaining: dict[tuple[str, str], Any] = {
        _merge_key(r.host, _recorder_kind(r)): r
        for r in existing
        if _recorder_kind(r) in CMDB_MANAGED_KINDS
    }
    old_keys = set(remaining.keys())

    merged_cmdb: list[Any] = []
    errors: list[MergeError] = []
    preserved = 0
    added = 0

    for row in cmdb_rows:
        try:
            base = {
                "object_name": row.object_name,
                "name": row.name,
                "host": row.host,
                "mac": row.mac,
                "device_kind": row.device_kind,
                "port": 80,
                "use_https": False,
            }
            key = _merge_key(row.host, row.device_kind)
            old = remaining.pop(key, None)
            if old:
                recorder = Recorder(
                    id=old.id,
                    last_status=old.last_status,
                    last_check_at=old.last_check_at,
                    last_error=old.last_error,
                    **base,
                )
                preserved += 1
            else:
                recorder = Recorder(id=new_device_id(row.device_kind), **base)
                added += 1
            merged_cmdb.append(recorder)
        except ValidationError as e:
            errors.append(MergeError(source_row=row.source_row, message=str(e)))

    merged = preserved_non_cmdb + merged_cmdb
    new_keys = {_merge_key(r.host, _recorder_kind(r)) for r in merged_cmdb}
    removed = len(old_keys - new_keys)
    stats = MergeStats(preserved=preserved, added=added, removed=removed)
    return merged, stats, errors


def merge_recorders_from_cmdb(
    cmdb_rows: list[CmdbDeviceRow],
    existing: list[Any],
) -> tuple[list[Any], MergeStats, list[MergeError]]:
    return merge_devices_from_cmdb(cmdb_rows, existing)
