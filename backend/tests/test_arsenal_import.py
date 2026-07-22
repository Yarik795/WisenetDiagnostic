"""Тесты импорта данных АС Арсенал."""

from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import Workbook

from app.arsenal_import import (
    COL_OBJECT_TYPE,
    COL_PASSPORT,
    SHEET_ANALYTICS,
    SHEET_GENERAL,
    import_arsenal_xlsx,
    normalize_manufacturer,
    parse_percent,
    parse_presence,
)
from app.config_store import ConfigStore
from app.data_sources import (
    ARSENAL_SOURCE,
    RunnerDeps,
    copy_to_storage,
    find_latest_source_file,
    load_source,
)
from app.state_store import StateStore
from app.ui.arsenal_dashboard import arsenal_dashboard_context

ANALYTICS_HEADERS = (
    "Наименование ТБ",
    "Наименование ГОСБ",
    COL_PASSPORT,
    "Статус паспорта",
    COL_OBJECT_TYPE,
    "Подтип объекта охраны",
    "Полное наименование объекта банка",
    "% заполнения паспорта",
    "% заполнения Основные сведения",
    "% заполнения Тех укрепленность",
    "% заполнения Посты",
    "% заполнения Договоры ФО",
    "% заполнения САЗ",
    "% заполнения СОУЭ",
    "% заполнения САПС",
    "% заполнения СОТС",
    "% заполнения ТСВ",
    "% заполнения СКУД",
    "Количество ошибок в паспорте",
    "Количество ошибок в Основные сведения",
    "Количество ошибок в Тех укрепленность",
    "Количество ошибок в Посты",
    "Количество ошибок в Договоры ФО",
    "Количество ошибок в САЗ",
    "Количество ошибок в СОУЭ",
    "Количество ошибок в САПС",
    "Количество ошибок в СОТС",
    "Количество ошибок в ТСВ",
    "Количество ошибок в СКУД",
    "% заполнения проектной документации",
    "Наличие документации в САЗ",
    "Наличие документации в СОУЭ",
    "Наличие документации в СОТС",
    "Наличие документации в САПС",
    "Наличие документации в ТСВ",
    "Наличие документации в СКУД",
    "Наличие прикрепленных фотографий в разделе Основная информация",
)


def _write_arsenal_xlsx(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_ANALYTICS
    ws.append(list(ANALYTICS_HEADERS))
    ws.append(
        [
            "Тестовый банк",
            "ОПЕРУ Тестовый банк",
            114585,
            "Действительный",
            "ВСП",
            "Ритейл",
            "Доп.офис №9038/030",
            100,
            100,
            100,
            100,
            100,
            100,
            100,
            100,
            100,
            100,
            100,
            0,
            0,
            0,
            0,
            0,
            0,
            0,
            0,
            0,
            0,
            0,
            "83,33333",
            "Да",
            "Да",
            "Да",
            "Да",
            "Да",
            "Да",
            "Нет",
        ]
    )

    tsv = wb.create_sheet("ТСВ")
    tsv.append(
        [
            "Наименование ТБ",
            "Наименование ГОСБ",
            COL_PASSPORT,
            "Статус паспорта",
            COL_OBJECT_TYPE,
            "Подтип объекта охраны",
            "Полное наименование объекта банка",
            "Наличие ТСВ",
            "Год установки, капитального ремонта ТСВ",
            "Количество аналоговых видеокамер камер на объекте",
            "Количество IP видеокамер камер на объекте",
            "Количество видеорегистраторов на объекте",
            "Вендоры IP оборудования на объекте",
        ]
    )
    tsv.append(
        [
            "Тестовый банк",
            "ОПЕРУ Тестовый банк",
            114585,
            "Действительный",
            "ВСП",
            "Ритейл",
            "Доп.офис №9038/030",
            "ДА",
            2025,
            29,
            2,
            3,
            "HANWHA",
        ]
    )

    skud = wb.create_sheet("СКУД")
    skud.append(
        [
            "Наименование ТБ",
            "Наименование ГОСБ",
            COL_PASSPORT,
            "Статус паспорта",
            COL_OBJECT_TYPE,
            "Подтип объекта охраны",
            "Полное наименование объекта банка",
            "Наличие СКУД",
            "Год установки, капитального ремонта СКУД",
            "Производитель контроллера",
            'Производитель (в случае если выбрано "Иное")',
        ]
    )
    skud.append(
        [
            "Тестовый банк",
            "ОПЕРУ Тестовый банк",
            114585,
            "Действительный",
            "ВСП",
            "Ритейл",
            "Доп.офис №9038/030",
            "ДА",
            2025,
            "Иное",
            "Бастион",
        ]
    )

    general = wb.create_sheet(SHEET_GENERAL)
    general.append(
        [
            "Наименование ТБ",
            "Наименование ГОСБ",
            COL_PASSPORT,
            "Статус паспорта",
            COL_OBJECT_TYPE,
            "Подтип объекта охраны",
            "Полное наименование объекта банка",
            "Адрес",
            "Уточненный фактический адрес расположения",
        ]
    )
    general.append(
        [
            "Тестовый банк",
            "ОПЕРУ Тестовый банк",
            114585,
            "Действительный",
            "ВСП",
            "Ритейл",
            "Доп.офис №9038/030",
            "г.Москва, ул.Тестовая, 1",
            "Россия, Москва, ул. Ясеневая, 26",
        ]
    )

    wb.save(path)


def _write_arsenal_xlsx_skud_absent(path: Path) -> None:
    """Паспорт с «Наличие СКУД» = НЕТ и пустым производителем."""
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_ANALYTICS
    ws.append(list(ANALYTICS_HEADERS))
    ws.append(
        [
            "Тестовый банк",
            "ОПЕРУ Тестовый банк",
            999001,
            "Действительный",
            "ВСП",
            "Ритейл",
            "Доп.офис без СКУД",
            100,
            100,
            100,
            100,
            100,
            100,
            100,
            100,
            100,
            100,
            100,
            0,
            0,
            0,
            0,
            0,
            0,
            0,
            0,
            0,
            0,
            0,
            "100",
            "Да",
            "Да",
            "Да",
            "Да",
            "Да",
            "Нет",
            "Нет",
        ]
    )

    skud = wb.create_sheet("СКУД")
    skud.append(
        [
            "Наименование ТБ",
            "Наименование ГОСБ",
            COL_PASSPORT,
            "Статус паспорта",
            COL_OBJECT_TYPE,
            "Подтип объекта охраны",
            "Полное наименование объекта банка",
            "Наличие СКУД",
            "Год установки, капитального ремонта СКУД",
            "Производитель контроллера",
            'Производитель (в случае если выбрано "Иное")',
        ]
    )
    skud.append(
        [
            "Тестовый банк",
            "ОПЕРУ Тестовый банк",
            999001,
            "Действительный",
            "ВСП",
            "Ритейл",
            "Доп.офис без СКУД",
            "НЕТ",
            None,
            "",
            "",
        ]
    )

    wb.save(path)


def test_parse_presence() -> None:
    assert parse_presence("ДА") is True
    assert parse_presence("да") is True
    assert parse_presence("НЕТ") is False
    assert parse_presence("нет") is False
    assert parse_presence("no") is False
    assert parse_presence("") is True
    assert parse_presence(None) is True
    assert parse_percent("83,33333") == pytest.approx(83.33333)
    assert parse_percent(100) == 100.0


def test_normalize_manufacturer_other() -> None:
    assert normalize_manufacturer("Иное", "ROXTON") == "ROXTON"
    assert normalize_manufacturer("HANWHA", "") == "HANWHA"
    assert normalize_manufacturer("", "") == "Не указан"


def test_import_arsenal_xlsx_parses_rows(tmp_path: Path) -> None:
    xlsx = tmp_path / "паспортам_выгрузка.xlsx"
    _write_arsenal_xlsx(xlsx)

    state = StateStore(path=tmp_path / "monitoring.db")
    state.init_db()

    count = import_arsenal_xlsx(xlsx, state)
    assert count == 1
    assert state.count_arsenal_records() == 1

    analytics = state.arsenal_analytics_rows()
    assert analytics[0].passport_number == "114585"
    assert analytics[0].object_type == "ВСП"
    assert analytics[0].fill_total == 100.0
    assert analytics[0].address == "Россия, Москва, ул. Ясеневая, 26"

    systems = state.arsenal_systems_rows()
    manufacturers = {row.system_type: row.manufacturer for row in systems}
    assert manufacturers["ТСВ"] == "HANWHA"
    assert manufacturers["СКУД"] == "Бастион"
    present_by_type = {row.system_type: row.present for row in systems}
    assert present_by_type["ТСВ"] is True
    assert present_by_type["СКУД"] is True


def test_import_skud_absent_excluded_from_manufacturer_chart(tmp_path: Path) -> None:
    xlsx = tmp_path / "паспортам_absent_skud.xlsx"
    _write_arsenal_xlsx_skud_absent(xlsx)

    state = StateStore(path=tmp_path / "monitoring.db")
    state.init_db()
    import_arsenal_xlsx(xlsx, state)

    systems = state.arsenal_systems_rows()
    skud_rows = [row for row in systems if row.system_type == "СКУД"]
    assert len(skud_rows) == 1
    assert skud_rows[0].present is False
    assert skud_rows[0].manufacturer == "Не указан"

    analytics = state.arsenal_analytics_rows()
    ctx = arsenal_dashboard_context(analytics, systems)
    skud_chart = ctx["arsenal_charts"]["systems"]["СКУД"]
    assert skud_chart["total"] == 0
    assert "Не указан" not in skud_chart["manufacturers"]


def test_arsenal_export_context(tmp_path: Path) -> None:
    xlsx = tmp_path / "паспортам_export.xlsx"
    _write_arsenal_xlsx(xlsx)

    state = StateStore(path=tmp_path / "monitoring.db")
    state.init_db()
    import_arsenal_xlsx(xlsx, state)

    from app.ui.arsenal_dashboard import arsenal_page_context
    from app.ui.arsenal_export import (
        build_arsenal_export_context,
        render_arsenal_export_html,
    )

    page_ctx = arsenal_page_context(state)
    export_ctx = build_arsenal_export_context(page_ctx)
    assert export_ctx["filter_label"] == "Все типы"
    assert export_ctx["kpi"]["passport_count"] == 1
    assert len(export_ctx["system_sections"]) == 6

    html = render_arsenal_export_html(export_ctx)
    assert "Арсенал" in html
    assert "<svg" in html
    assert "HANWHA" in html


def test_find_latest_source_file_arsenal_marker(tmp_path: Path) -> None:
    wrong = tmp_path / "other.xlsx"
    right = tmp_path / "паспортам_2025.xlsx"
    _write_arsenal_xlsx(wrong)
    _write_arsenal_xlsx(right)

    found = find_latest_source_file(ARSENAL_SOURCE, tmp_path)
    assert found == right


def test_load_source_arsenal_success(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    input_dir = tmp_path / "inputData"
    uploads = tmp_path / "uploads"
    input_dir.mkdir()
    uploads.mkdir()
    monkeypatch.setattr("app.data_sources.INPUT_DATA_DIR", input_dir)
    monkeypatch.setattr("app.data_sources.UPLOADS_DIR", uploads)

    source = input_dir / "паспортам_2025.xlsx"
    _write_arsenal_xlsx(source)

    store = ConfigStore(path=tmp_path / "config.json")
    state = StateStore(path=tmp_path / "monitoring.db")
    state.init_db()
    deps = RunnerDeps(store=store, state=state)

    result = load_source("arsenal", deps)
    assert result.ok
    assert result.record_count == 1
    assert "1 паспорт" in result.message
    assert state.count_arsenal_records() == 1


def test_load_source_arsenal_skip_unchanged(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    input_dir = tmp_path / "inputData"
    uploads = tmp_path / "uploads"
    input_dir.mkdir()
    uploads.mkdir()
    monkeypatch.setattr("app.data_sources.INPUT_DATA_DIR", input_dir)
    monkeypatch.setattr("app.data_sources.UPLOADS_DIR", uploads)

    source = input_dir / "паспортам_2025.xlsx"
    _write_arsenal_xlsx(source)
    copy_to_storage(ARSENAL_SOURCE, source)

    state = StateStore(path=tmp_path / "monitoring.db")
    state.init_db()
    import_arsenal_xlsx(copy_to_storage(ARSENAL_SOURCE, source), state)

    store = ConfigStore(path=tmp_path / "config.json")
    deps = RunnerDeps(store=store, state=state)

    result = load_source("arsenal", deps)
    assert result.ok
    assert result.message == "Новых данных нет"
    assert not result.changed
