"""Тесты отчёта «Анализ повторных РВР»."""

from __future__ import annotations

from datetime import date
from io import BytesIO

import pytest
from openpyxl import load_workbook

from app.rvr_repeat_report import (
    OBJECT_TYPE_ADZ,
    OBJECT_TYPE_VSP,
    build_rvr_repeat_report,
    extract_vk_comment,
    object_type_from_fio,
)
from app.ui.rvr_repeat_dashboard import filter_report_by_object_type
from app.ui.rvr_repeat_export import build_rvr_repeat_xlsx


def _row(
    *,
    request_number: str = "1001",
    status: str = "Отправлена",
    drug_number: str = "SD001",
    created_at: str = "2026-01-15T10:00:00+00:00",
    customer_fio: str = "Иванов",
    tb: str = "3800 Московский банк",
    work_type: str = "РВР",
    address: str = "ул. Ленина, 1",
    security_system_type: str = "СОТС",
) -> dict:
    return {
        "request_number": request_number,
        "status": status,
        "drug_number": drug_number,
        "created_at": created_at,
        "completed_at": None,
        "customer_fio": customer_fio,
        "tb": tb,
        "work_type": work_type,
        "act_status": "",
        "amount_vat": 0.0,
        "warranty": "Нет",
        "address": address,
        "security_system_type": security_system_type,
        "in_limit": "-",
        "raw_json": "{}",
        "source_row": 1,
    }


def test_extract_vk_comment() -> None:
    desc = "Префикс Комментарий ВК: не работает камера ФИО ВК хвост"
    assert extract_vk_comment(desc) == "не работает камера"
    assert extract_vk_comment("") == ""
    assert extract_vk_comment("без маркеров") == ""


def test_object_type_from_fio() -> None:
    assert object_type_from_fio("Олег Юрьевич К") == OBJECT_TYPE_ADZ
    assert object_type_from_fio("Случайный заказчик") == OBJECT_TYPE_VSP


def test_filters_exclude_non_rvr_and_revoked() -> None:
    rows = [
        _row(request_number="1", work_type="Модернизация"),
        _row(request_number="2", status="Отозвана ДБ"),
        _row(request_number="3", security_system_type="УС"),
        _row(request_number="4", tb="9900 ЦА"),
        _row(request_number="5", created_at="2025-12-31T23:59:00+00:00"),
    ]
    report = build_rvr_repeat_report(
        rows,
        {},
        date_from=date(2026, 1, 1),
        date_to=date(2026, 3, 31),
    )
    assert report["has_data"] is False
    assert report["data_rows"] == []


def test_repeat_count_and_thresholds() -> None:
    desc_map = {
        "SD001": "Комментарий ВК: проблема 1 ФИО ВК",
        "SD002": "Комментарий ВК: проблема 2 ФИО ВК",
        "SD003": "Комментарий ВК: проблема 3 ФИО ВК",
    }
    rows = [
        _row(request_number="1001", drug_number="SD001", created_at="2026-01-10T10:00:00+00:00"),
        _row(request_number="1002", drug_number="SD002", created_at="2026-02-10T10:00:00+00:00"),
        _row(request_number="1003", drug_number="SD003", created_at="2026-03-10T10:00:00+00:00"),
    ]
    report = build_rvr_repeat_report(
        rows,
        desc_map,
        date_from=date(2026, 1, 1),
        date_to=date(2026, 3, 31),
    )
    assert report["has_data"] is True
    assert len(report["data_rows"]) == 3
    assert len(report["groups_ge2"]) == 1
    assert report["groups_ge2"][0]["repeat_count"] == 2
    assert len(report["groups_ge3"]) == 1
    assert report["groups_ge3"][0]["repeat_count"] == 2
    assert report["kpi"]["groups_total"] == 1
    assert report["kpi"]["repeats_total"] == 2
    assert "ул. Ленина, 1" in report["kpi"]["top_object"]


def test_groups_sorted_by_repeats_desc() -> None:
    rows = [
        _row(
            request_number="a1",
            address="Адрес А",
            security_system_type="СОТС",
            created_at="2026-01-01T10:00:00+00:00",
        ),
        _row(
            request_number="a2",
            address="Адрес А",
            security_system_type="СОТС",
            created_at="2026-01-02T10:00:00+00:00",
        ),
        _row(
            request_number="b1",
            address="Адрес Б",
            security_system_type="ТСВ",
            created_at="2026-01-03T10:00:00+00:00",
        ),
        _row(
            request_number="b2",
            address="Адрес Б",
            security_system_type="ТСВ",
            created_at="2026-01-04T10:00:00+00:00",
        ),
        _row(
            request_number="b3",
            address="Адрес Б",
            security_system_type="СОТС",
            created_at="2026-01-05T10:00:00+00:00",
        ),
        _row(
            request_number="b4",
            address="Адрес Б",
            security_system_type="СОТС",
            created_at="2026-01-06T10:00:00+00:00",
        ),
    ]
    report = build_rvr_repeat_report(
        rows,
        {},
        date_from=date(2026, 1, 1),
        date_to=date(2026, 3, 31),
    )
    assert report["groups_ge2"][0]["address"] == "Адрес Б"
    assert report["groups_ge2"][0]["repeat_count"] == 2
    assert report["groups_ge2"][1]["address"] == "Адрес А"
    assert report["groups_ge2"][1]["repeat_count"] == 1


def test_build_rvr_repeat_xlsx_sheets() -> None:
    rows = [
        _row(request_number="1001"),
        _row(request_number="1002", created_at="2026-02-01T10:00:00+00:00"),
    ]
    report = build_rvr_repeat_report(
        rows,
        {},
        date_from=date(2026, 1, 1),
        date_to=date(2026, 3, 31),
    )
    xlsx = build_rvr_repeat_xlsx(report)
    wb = load_workbook(BytesIO(xlsx))
    assert wb.sheetnames == ["Данные", "Сводка", "Сводка 3"]
    ws_summary = wb["Сводка"]
    headers = [cell.value for cell in ws_summary[1]]
    assert "Адрес" in headers
    assert "Количество повторов" in headers
    assert "Анализ заявок / подозрение на повтор" in headers


def test_filter_report_by_object_type() -> None:
    report = {
        "groups_ge2": [
            {"address": "A", "object_type": OBJECT_TYPE_VSP, "repeat_count": 2},
            {"address": "B", "object_type": OBJECT_TYPE_ADZ, "repeat_count": 1},
        ],
        "groups_ge3": [
            {"address": "A", "object_type": OBJECT_TYPE_VSP, "repeat_count": 2},
        ],
        "data_rows": [
            {"object_type": OBJECT_TYPE_VSP},
            {"object_type": OBJECT_TYPE_VSP},
            {"object_type": OBJECT_TYPE_ADZ},
        ],
        "has_data": True,
        "kpi": {"groups_total": 2, "repeats_total": 3, "top_object": "A", "requests_total": 3},
    }
    filtered = filter_report_by_object_type(report, OBJECT_TYPE_VSP)
    assert len(filtered["groups_ge2"]) == 1
    assert filtered["groups_ge2"][0]["address"] == "A"
    assert len(filtered["data_rows"]) == 2
    assert filtered["kpi"]["groups_total"] == 1
    assert filtered["kpi"]["repeats_total"] == 2
    assert filtered["kpi"]["requests_total"] == 2
