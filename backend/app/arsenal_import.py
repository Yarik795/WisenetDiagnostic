"""Импорт выгрузки АС Арсенал (паспорта) из Excel в SQLite."""

from __future__ import annotations

import re
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import TYPE_CHECKING, Any, Callable, Optional

if TYPE_CHECKING:
    from .state_store import StateStore

ProgressCallback = Callable[[str, int], None]

SHEET_ANALYTICS = "Аналитика"

COL_TB = "Наименование ТБ"
COL_GOSB = "Наименование ГОСБ"
COL_PASSPORT = "Номер паспорта"
COL_STATUS = "Статус паспорта"
COL_OBJECT_TYPE = "Тип объекта охраны"
COL_SUBTYPE = "Подтип объекта охраны"
COL_OBJECT_NAME = "Полное наименование объекта банка"

COL_FILL_TOTAL = "% заполнения паспорта"
COL_FILL_PROJECT = "% заполнения проектной документации"
COL_ERRORS_TOTAL = "Количество ошибок в паспорте"
COL_HAS_PHOTOS = "Наличие прикрепленных фотографий в разделе Основная информация"

FILL_SECTIONS: tuple[tuple[str, str], ...] = (
    ("Основные сведения", "% заполнения Основные сведения"),
    ("Тех укрепленность", "% заполнения Тех укрепленность"),
    ("Посты", "% заполнения Посты"),
    ("Договоры ФО", "% заполнения Договоры ФО"),
    ("САЗ", "% заполнения САЗ"),
    ("СОУЭ", "% заполнения СОУЭ"),
    ("САПС", "% заполнения САПС"),
    ("СОТС", "% заполнения СОТС"),
    ("ТСВ", "% заполнения ТСВ"),
    ("СКУД", "% заполнения СКУД"),
)

ERROR_SECTIONS: tuple[tuple[str, str], ...] = (
    ("Основные сведения", "Количество ошибок в Основные сведения"),
    ("Тех укрепленность", "Количество ошибок в Тех укрепленность"),
    ("Посты", "Количество ошибок в Посты"),
    ("Договоры ФО", "Количество ошибок в Договоры ФО"),
    ("САЗ", "Количество ошибок в САЗ"),
    ("СОУЭ", "Количество ошибок в СОУЭ"),
    ("САПС", "Количество ошибок в САПС"),
    ("СОТС", "Количество ошибок в СОТС"),
    ("ТСВ", "Количество ошибок в ТСВ"),
    ("СКУД", "Количество ошибок в СКУД"),
)

DOC_COLUMNS: tuple[tuple[str, str], ...] = (
    ("САЗ", "Наличие документации в САЗ"),
    ("СОУЭ", "Наличие документации в СОУЭ"),
    ("СОТС", "Наличие документации в СОТС"),
    ("САПС", "Наличие документации в САПС"),
    ("ТСВ", "Наличие документации в ТСВ"),
    ("СКУД", "Наличие документации в СКУД"),
)

ANALYTICS_REQUIRED_COLUMNS = (
    COL_TB,
    COL_GOSB,
    COL_PASSPORT,
    COL_STATUS,
    COL_OBJECT_TYPE,
    COL_SUBTYPE,
    COL_OBJECT_NAME,
    COL_FILL_TOTAL,
    COL_ERRORS_TOTAL,
)

COMMON_COLUMNS = (
    COL_TB,
    COL_GOSB,
    COL_PASSPORT,
    COL_STATUS,
    COL_OBJECT_TYPE,
    COL_SUBTYPE,
    COL_OBJECT_NAME,
)

BATCH_SIZE = 500

_OTHER_MARKERS = frozenset({"иное", "другое", "-", "—"})


@dataclass(frozen=True)
class SystemSheetSpec:
    sheet_name: str
    system_type: str
    manufacturer_col: str
    manufacturer_other_col: Optional[str] = None
    year_col: Optional[str] = None


SYSTEM_SHEETS: tuple[SystemSheetSpec, ...] = (
    SystemSheetSpec(
        sheet_name="САЗ",
        system_type="САЗ",
        manufacturer_col="Производитель",
        manufacturer_other_col="Производитель (Иной)",
        year_col="Год установки / капитального ремонта САЗ",
    ),
    SystemSheetSpec(
        sheet_name="СОУЭ",
        system_type="СОУЭ",
        manufacturer_col="Производитель блоков управления",
        manufacturer_other_col="Производитель блоков управления (Иной)",
        year_col="Год установки / капитального ремонта СОУЭ",
    ),
    SystemSheetSpec(
        sheet_name="СОТС",
        system_type="СОТС",
        manufacturer_col="Производитель объектовой сигнализации",
        manufacturer_other_col="Производитель (Иной)",
        year_col="Год установки, капитального ремонта СOTC",
    ),
    SystemSheetSpec(
        sheet_name="САПС",
        system_type="САПС",
        manufacturer_col="Производитель",
        manufacturer_other_col="Производитель (Иной)",
        year_col="Год установки/капитального ремонта САПС",
    ),
    SystemSheetSpec(
        sheet_name="ТСВ",
        system_type="ТСВ",
        manufacturer_col="Вендоры IP оборудования на объекте",
        year_col="Год установки, капитального ремонта ТСВ",
    ),
    SystemSheetSpec(
        sheet_name="СКУД",
        system_type="СКУД",
        manufacturer_col="Производитель контроллера",
        manufacturer_other_col='Производитель (в случае если выбрано "Иное")',
        year_col="Год установки, капитального ремонта СКУД",
    ),
)


def _cell_str(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _normalize_header(value: Any) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value).strip())


def parse_percent(value: Any) -> float:
    if value is None or value == "":
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = re.sub(r"\s+", "", str(value).strip()).replace(",", ".")
    if not text:
        return 0.0
    try:
        return float(text)
    except ValueError:
        return 0.0


def parse_int(value: Any) -> int:
    if value is None or value == "":
        return 0
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        return int(value)
    text = re.sub(r"\s+", "", str(value).strip()).replace(",", ".")
    if not text:
        return 0
    try:
        return int(float(text))
    except ValueError:
        return 0


def parse_year(value: Any) -> Optional[int]:
    if value is None or value == "":
        return None
    if isinstance(value, int):
        return value if 1900 <= value <= 2100 else None
    if isinstance(value, float):
        year = int(value)
        return year if 1900 <= year <= 2100 else None
    text = re.sub(r"\s+", "", str(value).strip())
    if not text:
        return None
    try:
        year = int(float(text.replace(",", ".")))
    except ValueError:
        return None
    return year if 1900 <= year <= 2100 else None


def build_col_index(
    header_row: list[Any],
    required: tuple[str, ...],
) -> dict[str, int]:
    index: dict[str, int] = {}
    for col, cell in enumerate(header_row):
        name = _normalize_header(cell)
        if name and name not in index:
            index[name] = col
    missing = [col for col in required if col not in index]
    if missing:
        raise ValueError(
            "В заголовке отсутствуют колонки: " + ", ".join(missing)
        )
    return index


def normalize_manufacturer(primary: str, other: str = "") -> str:
    primary_clean = primary.strip()
    other_clean = other.strip()
    primary_lower = primary_clean.lower()
    if not primary_clean or primary_lower in _OTHER_MARKERS:
        return other_clean or "Не указан"
    if primary_lower == "иное" and other_clean:
        return other_clean
    return primary_clean


def _value_at(values: tuple[Any, ...], col_index: dict[str, int], column: str) -> Any:
    idx = col_index.get(column)
    if idx is None or idx >= len(values):
        return None
    return values[idx]


def _analytics_row_from_values(
    values: tuple[Any, ...],
    col_index: dict[str, int],
    source_row: int,
) -> Optional["ArsenalAnalyticsRow"]:
    from .state_store import ArsenalAnalyticsRow

    passport = _cell_str(_value_at(values, col_index, COL_PASSPORT))
    if not passport:
        return None

    fill_sections = {
        label: parse_percent(_value_at(values, col_index, col))
        for label, col in FILL_SECTIONS
        if col in col_index
    }
    errors_sections = {
        label: parse_int(_value_at(values, col_index, col))
        for label, col in ERROR_SECTIONS
        if col in col_index
    }
    docs = {
        label: _cell_str(_value_at(values, col_index, col)) or "-"
        for label, col in DOC_COLUMNS
        if col in col_index
    }

    return ArsenalAnalyticsRow(
        passport_number=passport,
        tb=_cell_str(_value_at(values, col_index, COL_TB)),
        gosb=_cell_str(_value_at(values, col_index, COL_GOSB)),
        status=_cell_str(_value_at(values, col_index, COL_STATUS)),
        object_type=_cell_str(_value_at(values, col_index, COL_OBJECT_TYPE)),
        subtype=_cell_str(_value_at(values, col_index, COL_SUBTYPE)),
        object_name=_cell_str(_value_at(values, col_index, COL_OBJECT_NAME)),
        fill_total=parse_percent(_value_at(values, col_index, COL_FILL_TOTAL)),
        fill_sections=fill_sections,
        errors_total=parse_int(_value_at(values, col_index, COL_ERRORS_TOTAL)),
        errors_sections=errors_sections,
        fill_project_docs=parse_percent(
            _value_at(values, col_index, COL_FILL_PROJECT)
            if COL_FILL_PROJECT in col_index
            else None
        ),
        docs=docs,
        has_photos=_cell_str(
            _value_at(values, col_index, COL_HAS_PHOTOS)
            if COL_HAS_PHOTOS in col_index
            else None
        ),
        source_row=source_row,
    )


def _system_row_from_values(
    values: tuple[Any, ...],
    col_index: dict[str, int],
    spec: SystemSheetSpec,
    source_row: int,
) -> Optional["ArsenalSystemRow"]:
    from .state_store import ArsenalSystemRow

    passport = _cell_str(_value_at(values, col_index, COL_PASSPORT))
    if not passport:
        return None

    primary = _cell_str(_value_at(values, col_index, spec.manufacturer_col))
    other = ""
    if spec.manufacturer_other_col:
        other = _cell_str(
            _value_at(values, col_index, spec.manufacturer_other_col)
        )
    manufacturer = normalize_manufacturer(primary, other)

    year = None
    if spec.year_col and spec.year_col in col_index:
        year = parse_year(_value_at(values, col_index, spec.year_col))

    return ArsenalSystemRow(
        passport_number=passport,
        tb=_cell_str(_value_at(values, col_index, COL_TB)),
        gosb=_cell_str(_value_at(values, col_index, COL_GOSB)),
        object_type=_cell_str(_value_at(values, col_index, COL_OBJECT_TYPE)),
        subtype=_cell_str(_value_at(values, col_index, COL_SUBTYPE)),
        system_type=spec.system_type,
        manufacturer=manufacturer,
        year=year,
        source_row=source_row,
    )


def _import_analytics_sheet(
    wb: Any,
    session: Any,
    on_progress: ProgressCallback,
) -> int:
    if SHEET_ANALYTICS not in wb.sheetnames:
        raise ValueError(f'В файле отсутствует лист «{SHEET_ANALYTICS}»')

    ws = wb[SHEET_ANALYTICS]
    row_iter = ws.iter_rows(values_only=True)
    try:
        header_row = list(next(row_iter))
    except StopIteration:
        raise ValueError("Лист «Аналитика» не содержит строк с данными") from None

    col_index = build_col_index(header_row, ANALYTICS_REQUIRED_COLUMNS)
    max_row = ws.max_row or 1
    data_rows = max(max_row - 1, 1)
    batch: list[Any] = []

    for row_idx, values in enumerate(row_iter, start=2):
        if values is None:
            continue
        row = _analytics_row_from_values(tuple(values), col_index, row_idx)
        if row is None:
            continue
        batch.append(row)
        if len(batch) >= BATCH_SIZE:
            session.write_analytics_batch(batch)
            batch = []
            if session.analytics_count % (BATCH_SIZE * 2) == 0:
                percent = min(45, 10 + int(session.analytics_count / data_rows * 35))
                on_progress("Чтение Аналитика", percent)

    if batch:
        session.write_analytics_batch(batch)

    return session.analytics_count


def _import_system_sheet(
    wb: Any,
    spec: SystemSheetSpec,
    session: Any,
    on_progress: ProgressCallback,
    *,
    sheet_index: int,
    sheet_total: int,
) -> None:
    if spec.sheet_name not in wb.sheetnames:
        return

    ws = wb[spec.sheet_name]
    row_iter = ws.iter_rows(values_only=True)
    try:
        header_row = list(next(row_iter))
    except StopIteration:
        return

    try:
        col_index = build_col_index(header_row, COMMON_COLUMNS + (spec.manufacturer_col,))
    except ValueError:
        return

    batch: list[Any] = []
    base_percent = 45 + int(sheet_index / sheet_total * 45)

    for row_idx, values in enumerate(row_iter, start=2):
        if values is None:
            continue
        row = _system_row_from_values(tuple(values), col_index, spec, row_idx)
        if row is None:
            continue
        batch.append(row)
        if len(batch) >= BATCH_SIZE:
            session.write_systems_batch(batch)
            batch = []

    if batch:
        session.write_systems_batch(batch)

    on_progress(f"Чтение {spec.system_type}", min(90, base_percent + 5))


def import_arsenal_xlsx(
    path: Path,
    state: "StateStore",
    on_progress: Optional[ProgressCallback] = None,
) -> int:
    """Читает xlsx и полностью заменяет данные Арсенал в SQLite."""
    from openpyxl import load_workbook

    progress = on_progress or (lambda _phase, _percent: None)
    imported_at = datetime.now(timezone.utc)

    progress("Чтение Арсенал", 5)
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        with state.replace_arsenal_data(imported_at) as session:
            count = _import_analytics_sheet(wb, session, progress)
            if count == 0:
                raise ValueError("Лист «Аналитика» не содержит валидных паспортов")

            sheet_total = len(SYSTEM_SHEETS)
            for idx, spec in enumerate(SYSTEM_SHEETS):
                _import_system_sheet(
                    wb,
                    spec,
                    session,
                    progress,
                    sheet_index=idx,
                    sheet_total=sheet_total,
                )

        progress("Сохранение", 95)
        return count
    finally:
        wb.close()
