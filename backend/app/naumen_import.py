"""Импорт выгрузки Naumen из Excel в SQLite."""

from __future__ import annotations

import re
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import TYPE_CHECKING, Any, Callable, Optional

if TYPE_CHECKING:
    from .state_store import StateStore

ProgressCallback = Callable[[str, int], None]

COL_EXTERNAL_ID = "ID внешней системы"
COL_NUMBER = "Номер"
COL_COST = "Стоимость"
COL_SBERDRUG = "Номер Сбердруг"
COL_DESCRIPTION = "Описание"

REQUIRED_COLUMNS = (
    COL_EXTERNAL_ID,
    COL_NUMBER,
    COL_COST,
    COL_SBERDRUG,
    COL_DESCRIPTION,
)

BATCH_SIZE = 500


@dataclass(frozen=True)
class NaumenRow:
    external_id: str
    number: str
    cost: float
    sberdrug_number: str
    description: str
    source_row: int


def _cell_str(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def parse_cost(value: Any) -> float:
    if value is None or value == "":
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = re.sub(r"\s+", "", str(value).strip()).replace(",", ".")
    if not text:
        return 0.0
    try:
        return float(text)
    except ValueError:
        return 0.0


def _normalize_header(value: Any) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value).strip())


def build_col_index(header_row: list[Any]) -> dict[str, int]:
    index: dict[str, int] = {}
    for col, cell in enumerate(header_row):
        name = _normalize_header(cell)
        if name and name not in index:
            index[name] = col
    missing = [col for col in REQUIRED_COLUMNS if col not in index]
    if missing:
        raise ValueError(
            "В заголовке отсутствуют колонки: " + ", ".join(missing)
        )
    return index


def _row_from_values(
    values: tuple[Any, ...],
    col_index: dict[str, int],
    source_row: int,
) -> Optional[NaumenRow]:
    external_id = _cell_str(values[col_index[COL_EXTERNAL_ID]])
    if not external_id:
        return None
    return NaumenRow(
        external_id=external_id,
        number=_cell_str(values[col_index[COL_NUMBER]]),
        cost=parse_cost(values[col_index[COL_COST]] if col_index[COL_COST] < len(values) else None),
        sberdrug_number=_cell_str(
            values[col_index[COL_SBERDRUG]] if col_index[COL_SBERDRUG] < len(values) else None
        ),
        description=_cell_str(
            values[col_index[COL_DESCRIPTION]] if col_index[COL_DESCRIPTION] < len(values) else None
        ),
        source_row=source_row,
    )


def import_naumen_xlsx(
    path: Path,
    state: "StateStore",
    on_progress: Optional[ProgressCallback] = None,
) -> int:
    """Читает xlsx потоково и полностью заменяет naumen_records в SQLite."""
    from openpyxl import load_workbook

    progress = on_progress or (lambda _phase, _percent: None)
    imported_at = datetime.now(timezone.utc)

    progress("Чтение Naumen", 30)

    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb.active
        row_iter = ws.iter_rows(values_only=True)
        try:
            header_row = list(next(row_iter))
        except StopIteration:
            raise ValueError("Файл не содержит строк с данными") from None

        col_index = build_col_index(header_row)
        max_row = ws.max_row or 1
        data_rows = max(max_row - 1, 1)

        with state.replace_naumen_records(imported_at) as session:
            batch: list[NaumenRow] = []

            for row_idx, values in enumerate(row_iter, start=2):
                if values is None:
                    continue
                row = _row_from_values(tuple(values), col_index, row_idx)
                if row is None:
                    continue
                batch.append(row)
                if len(batch) >= BATCH_SIZE:
                    session.write_batch(batch)
                    batch = []
                    if session.count % (BATCH_SIZE * 4) == 0:
                        percent = min(85, 30 + int(session.count / data_rows * 55))
                        progress("Чтение Naumen", percent)

            if batch:
                session.write_batch(batch)

        progress("Сохранение", 95)
        return session.count
    finally:
        wb.close()
