"""Импорт выгрузки заявок ПП из Excel в SQLite."""

from __future__ import annotations

import json
import logging
import re
from dataclasses import dataclass
from datetime import date, datetime, timezone
from pathlib import Path
from typing import TYPE_CHECKING, Any, Callable, Optional

if TYPE_CHECKING:
    from .state_store import StateStore

logger = logging.getLogger(__name__)

ProgressCallback = Callable[[str, int], None]

COL_REQUEST_NUMBER = "Заявка №"
COL_STATUS = "Статус"
COL_DRUG = "№ заявки ДРУГ"
COL_CREATED_AT = "Дата создания (UTC)"
COL_COMPLETED_AT = "Фактическая дата выполнения (UTC)"
COL_CUSTOMER_FIO = "ФИО заказчика"
COL_TB = "Территориальный банк"
COL_WORK_TYPE = "Вид работ"
COL_ACT_STATUS = "Статус акта"
COL_AMOUNT = "Сумма с НДС"
COL_WARRANTY = "Гарантийная заявка"
COL_ADDRESS = "Адрес"
COL_SECURITY_TYPE = "Вид системы безопасности"
COL_IN_LIMIT = "В лимите"

MIN_REQUIRED_COLUMNS = (
    COL_REQUEST_NUMBER,
    COL_AMOUNT,
    COL_COMPLETED_AT,
)

BATCH_SIZE = 500


@dataclass(frozen=True)
class PPRequestRow:
    request_number: str
    status: str
    drug_number: str
    created_at: Optional[str]
    completed_at: Optional[str]
    customer_fio: str
    tb: str
    work_type: str
    act_status: str
    amount_vat: float
    warranty: str
    address: str
    security_system_type: str
    in_limit: str
    raw_json: str
    source_row: int


def _cell_str(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def parse_money(value: Any) -> float:
    if value is None or value == "":
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = re.sub(r"\s+", "", str(value).strip()).replace(",", ".")
    if not text:
        return 0.0
    try:
        return float(text)
    except ValueError:
        return 0.0


def parse_dt(value: Any) -> Optional[str]:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        dt = value
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=timezone.utc)
        else:
            dt = dt.astimezone(timezone.utc)
        return dt.isoformat()
    if isinstance(value, date):
        return datetime.combine(value, datetime.min.time(), tzinfo=timezone.utc).isoformat()
    text = re.sub(r"\s+", " ", str(value).strip())
    if not text:
        return None
    for fmt in ("%d.%m.%Y %H:%M:%S", "%d.%m.%Y %H:%M"):
        try:
            parsed = datetime.strptime(text, fmt).replace(tzinfo=timezone.utc)
            return parsed.isoformat()
        except ValueError:
            continue
    return None


def _normalize_header(value: Any) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value).strip())


def _serialize_cell(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, datetime):
        return parse_dt(value)
    if isinstance(value, date):
        return parse_dt(value)
    if isinstance(value, (int, float)):
        return value
    text = str(value).strip()
    return text if text else None


def build_col_index(header_row: list[Any]) -> dict[str, int]:
    index: dict[str, int] = {}
    for col, cell in enumerate(header_row):
        name = _normalize_header(cell)
        if name and name not in index:
            index[name] = col
    missing = [col for col in MIN_REQUIRED_COLUMNS if col not in index]
    if missing:
        raise ValueError(
            "В заголовке отсутствуют колонки: " + ", ".join(missing)
        )
    return index


def _get_cell(values: tuple[Any, ...], col_index: dict[str, int], column: str) -> Any:
    idx = col_index.get(column)
    if idx is None or idx >= len(values):
        return None
    return values[idx]


def _build_raw_json(
    values: tuple[Any, ...],
    header_row: list[Any],
    col_index: dict[str, int],
) -> str:
    payload: dict[str, Any] = {}
    for name, idx in col_index.items():
        if idx < len(values):
            payload[name] = _serialize_cell(values[idx])
    return json.dumps(payload, ensure_ascii=False)


def _row_from_values(
    values: tuple[Any, ...],
    col_index: dict[str, int],
    header_row: list[Any],
    source_row: int,
) -> Optional[PPRequestRow]:
    request_number = _cell_str(_get_cell(values, col_index, COL_REQUEST_NUMBER))
    if not request_number:
        return None
    return PPRequestRow(
        request_number=request_number,
        status=_cell_str(_get_cell(values, col_index, COL_STATUS)),
        drug_number=_cell_str(_get_cell(values, col_index, COL_DRUG)),
        created_at=parse_dt(_get_cell(values, col_index, COL_CREATED_AT)),
        completed_at=parse_dt(_get_cell(values, col_index, COL_COMPLETED_AT)),
        customer_fio=_cell_str(_get_cell(values, col_index, COL_CUSTOMER_FIO)),
        tb=_cell_str(_get_cell(values, col_index, COL_TB)),
        work_type=_cell_str(_get_cell(values, col_index, COL_WORK_TYPE)),
        act_status=_cell_str(_get_cell(values, col_index, COL_ACT_STATUS)),
        amount_vat=parse_money(_get_cell(values, col_index, COL_AMOUNT)),
        warranty=_cell_str(_get_cell(values, col_index, COL_WARRANTY)),
        address=_cell_str(_get_cell(values, col_index, COL_ADDRESS)),
        security_system_type=_cell_str(
            _get_cell(values, col_index, COL_SECURITY_TYPE)
        ),
        in_limit=_cell_str(_get_cell(values, col_index, COL_IN_LIMIT)),
        raw_json=_build_raw_json(values, header_row, col_index),
        source_row=source_row,
    )


def import_pp_requests_xlsx(
    path: Path,
    state: "StateStore",
    on_progress: Optional[ProgressCallback] = None,
) -> int:
    """Читает xlsx потоково и полностью заменяет pp_requests в SQLite."""
    from openpyxl import load_workbook

    progress = on_progress or (lambda _phase, _percent: None)
    imported_at = datetime.now(timezone.utc)

    progress("Чтение заявок ПП", 30)

    wb = load_workbook(path, read_only=True, data_only=True)
    duplicate_numbers = 0
    seen_numbers: set[str] = set()
    try:
        ws = wb.active
        row_iter = ws.iter_rows(values_only=True)
        try:
            header_row = list(next(row_iter))
        except StopIteration:
            raise ValueError("Файл не содержит строк с данными") from None

        col_index = build_col_index(header_row)
        max_row = ws.max_row or 1
        data_rows = max(max_row - 1, 1)

        with state.replace_pp_requests(imported_at) as session:
            batch: list[PPRequestRow] = []

            for row_idx, values in enumerate(row_iter, start=2):
                if values is None:
                    continue
                row = _row_from_values(tuple(values), col_index, header_row, row_idx)
                if row is None:
                    continue
                if row.request_number in seen_numbers:
                    duplicate_numbers += 1
                seen_numbers.add(row.request_number)
                batch.append(row)
                if len(batch) >= BATCH_SIZE:
                    session.write_batch(batch)
                    batch = []
                    if session.count % (BATCH_SIZE * 4) == 0:
                        percent = min(85, 30 + int(session.count / data_rows * 55))
                        progress("Чтение заявок ПП", percent)

            if batch:
                session.write_batch(batch)

        if duplicate_numbers:
            logger.warning(
                "pp_import: %d duplicate request_number values replaced (last row wins)",
                duplicate_numbers,
            )

        progress("Сохранение", 95)
        return session.count
    finally:
        wb.close()
