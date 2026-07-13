"""XLSX-экспорт и email отчёта «Анализ повторных РВР»."""

from __future__ import annotations

from datetime import datetime, timezone
from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.worksheet.worksheet import Worksheet

from ..display_time import format_for_display
from ..rvr_ai_analysis import verdict_label
from ..rvr_repeat_report import format_kind_cell

ANALYSIS_COLUMN = "Анализ заявок / подозрение на повтор"
DESCRIPTION_COLUMN = "Описание проблем на объекте"
VERDICT_AI_COLUMN = "Вердикт AI"
XLSX_MIME = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)

_DATA_COLUMNS = (
    ("request_number", "Заявка №"),
    ("status", "Статус"),
    ("drug_number", "№ заявки ДРУГ"),
    ("created_at", "Дата создания (UTC)"),
    ("completed_at", "Фактическая дата выполнения (UTC)"),
    ("customer_fio", "ФИО заказчика"),
    ("tb", "Территориальный банк"),
    ("work_type", "Вид работ"),
    ("act_status", "Статус акта"),
    ("amount_vat", "Сумма с НДС"),
    ("warranty", "Гарантийная заявка"),
    ("address", "Адрес"),
    ("security_system_type", "Вид системы безопасности"),
    ("in_limit", "В лимите"),
    ("object_type", "Тип объекта"),
    ("description", "Описание"),
)


def _write_summary_sheet(
    ws: Worksheet,
    groups: list[dict[str, Any]],
    kinds: list[str],
    *,
    include_analysis_column: bool = True,
) -> None:
    wrap = Alignment(wrap_text=True, vertical="top")
    header = ["Адрес", "Тип объекта", "Количество повторов"] + kinds
    if include_analysis_column:
        header.append(VERDICT_AI_COLUMN)
        header.append(ANALYSIS_COLUMN)
        header.append(DESCRIPTION_COLUMN)
    ws.append(header)
    for group in groups:
        row: list[Any] = [
            group["address"],
            group["object_type"],
            group["repeat_count"],
        ]
        by_kind = group.get("by_kind") or {}
        for kind in kinds:
            entries = by_kind.get(kind, [])
            row.append(format_kind_cell(entries))
        if include_analysis_column:
            row.append(verdict_label(group.get("ai_verdict")))
            row.append(group.get("analysis") or "")
            row.append(group.get("description") or "")
        ws.append(row)
    for row_cells in ws.iter_rows(
        min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column
    ):
        for cell in row_cells:
            cell.alignment = wrap


def build_rvr_repeat_xlsx(
    report: dict[str, Any],
    *,
    include_analysis_column: bool = True,
) -> bytes:
    wb = Workbook()
    ws_data = wb.active
    ws_data.title = "Данные"
    ws_data.append([label for _, label in _DATA_COLUMNS])
    for row in report.get("data_rows") or []:
        ws_data.append([row.get(key, "") for key, _ in _DATA_COLUMNS])

    kinds = report.get("kinds") or []
    wrap = Alignment(wrap_text=True, vertical="top")
    for row_cells in ws_data.iter_rows(
        min_row=1, max_row=ws_data.max_row, min_col=1, max_col=ws_data.max_column
    ):
        for cell in row_cells:
            cell.alignment = wrap

    ws2 = wb.create_sheet("Сводка")
    _write_summary_sheet(
        ws2,
        report.get("groups_ge2") or [],
        kinds,
        include_analysis_column=include_analysis_column,
    )
    ws3 = wb.create_sheet("Сводка 3")
    _write_summary_sheet(
        ws3,
        report.get("groups_ge3") or [],
        kinds,
        include_analysis_column=include_analysis_column,
    )

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def rvr_repeat_export_filename(report: dict[str, Any]) -> str:
    period = report.get("period") or {}
    date_from = (period.get("from") or "start").replace("-", "")
    date_to = (period.get("to") or "end").replace("-", "")
    stamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    return f"rvr_repeat_{date_from}_{date_to}_{stamp}.xlsx"


def rvr_repeat_email_subject(report: dict[str, Any]) -> str:
    period = report.get("period") or {}
    label = period.get("label") or "период не указан"
    return f"Wisenet Диагностика — Анализ повторных РВР ({label})"


def render_rvr_repeat_email_body(report: dict[str, Any]) -> str:
    period = report.get("period") or {}
    kpi = report.get("kpi") or {}
    generated = report.get("generated_at")
    gen_display = ""
    if generated:
        try:
            gen_display = format_for_display(datetime.fromisoformat(generated))
        except ValueError:
            gen_display = generated

    confirmed_count = 0
    suspect_count = 0
    for group in report.get("groups_ge2") or []:
        verdict = group.get("ai_verdict")
        if verdict == "confirmed":
            confirmed_count += 1
        elif verdict in ("suspect", "repeat"):
            suspect_count += 1

    return f"""<!DOCTYPE html>
<html lang="ru">
<head><meta charset="utf-8"></head>
<body style="font-family:Segoe UI,Arial,sans-serif;color:#1e293b;">
  <h2>Анализ повторных РВР</h2>
  <p><strong>Период:</strong> {period.get("label", "—")}</p>
  <p><strong>Сформирован:</strong> {gen_display or "—"}</p>
  <ul>
    <li>Групп с повторами (≥2): {kpi.get("groups_total", 0)}</li>
    <li>Всего повторов: {kpi.get("repeats_total", 0)}</li>
    <li>Заявок в выборке: {kpi.get("requests_total", 0)}</li>
    <li>Топ-объект: {kpi.get("top_object") or "—"}</li>
    <li>Объектов с повтором (AI, подтверждён): {confirmed_count}</li>
    <li>Объектов с подозрением на повтор (AI): {suspect_count}</li>
  </ul>
  <p>Во вложении — файл Excel с листами «Данные», «Сводка» и «Сводка 3».</p>
</body>
</html>"""
